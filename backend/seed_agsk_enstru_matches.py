"""
Одноразовый скрипт: заполнение таблицы agsk_enstru_matches из Excel-файла.

Запуск:
    pip install openpyxl psycopg2-binary
    python seed_agsk_enstru_matches.py

Логика цветов:
  FF92D050  зелёный    — is_approved=True,  created_by/approved_by = аналитик блока
  FFFFC000  оранжевый  — блок аналитика 2  (строки Excel 2 – 74 299)
  theme:7   фиолетовый — блок аналитика 4  (строки Excel 74 300 – 147 174)
  theme:8   голубой    — блок аналитика 3  (строки Excel 147 175 – конец)

Зелёные строки утверждаются сразу (is_approved=True) с аналитиком своего блока.
Остальные — is_approved=False (ожидают утверждения).

ВАЖНО: одна ячейка F может содержать НЕСКОЛЬКО кодов ЕНСТРУ через запятую
(например "273213.700.000076, 273213.730.000079"). Каждый код выносится в
отдельную строку (agsk_code, enstru_code) — таблица описывает связь
многие-ко-многим (UNIQUE по паре agsk_code+enstru_code). Без этого длинная
строка из нескольких кодов не влезает в VARCHAR(35) и вставка падает с
ошибкой StringDataRightTruncation.
"""

import re
import sys
import logging
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ── Настройки БД ──────────────────────────────────────────────────────────────

DB_HOST     = "10.50.0.2"
DB_PORT     = 5432
DB_NAME     = "vc_system"
DB_USER     = "data_owner"
DB_PASSWORD = "DrVc2025"

# ── Путь к файлу ──────────────────────────────────────────────────────────────

EXCEL_PATH = Path(r"ВСЕ АГСК 08.06.2026.xlsx")

# ── ID аналитиков ─────────────────────────────────────────────────────────────

ANALYST_ORANGE = 2   # оранжевый блок  (строки 2 – 74 299)
ANALYST_PURPLE = 4   # фиолетовый блок (строки 74 300 – 147 174)
ANALYST_BLUE   = 3   # голубой блок    (строки 147 175 – конец)

# Строки начала каждого блока (нумерация Excel, 1-based, строка 1 = заголовок)
BLOCK_PURPLE_START = 74300
BLOCK_BLUE_START   = 147175

# Цвета
GREEN  = "FF92D050"

# Ограничение длины колонки enstru_code в БД (VARCHAR(35))
ENSTRU_MAX_LEN = 35

# ── Логирование ───────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s  %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("seed_agsk_enstru_matches.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ── Helpers ───────────────────────────────────────────────────────────────────

def row_color_key(cell) -> str:
    fc = cell.fill.fgColor
    if fc.type == "rgb":
        return fc.value
    if fc.type == "theme":
        return f"theme:{fc.theme}:{round(fc.tint, 2)}"
    return "none"


def analyst_for_row(excel_row_num: int) -> int:
    if excel_row_num >= BLOCK_BLUE_START:
        return ANALYST_BLUE
    if excel_row_num >= BLOCK_PURPLE_START:
        return ANALYST_PURPLE
    return ANALYST_ORANGE


def split_enstru_codes(raw: str) -> list[str]:
    """Разбивает содержимое ячейки F на отдельные коды ЕНСТРУ.

    Разделители: запятая и точка с запятой (с любыми пробелами вокруг).
    Возвращает список непустых обрезанных кодов.
    """
    return [c.strip() for c in re.split(r"[;,]", raw) if c.strip()]


# ── Основная логика ───────────────────────────────────────────────────────────

def main():
    log.info("Открываем файл: %s", EXCEL_PATH)
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb.active
    log.info("Строк в файле (с заголовком): %d", ws.max_row)

    log.info("Подключаемся к БД %s/%s…", DB_HOST, DB_NAME)
    conn = psycopg2.connect(
        host=DB_HOST, port=DB_PORT,
        dbname=DB_NAME, user=DB_USER, password=DB_PASSWORD,
    )
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM agsk_enstru_matches")
    existing = cur.fetchone()[0]
    if existing > 0:
        ans = input(f"\nВ таблице уже {existing} записей. Продолжить (дубли будут пропущены)? [y/N]: ")
        if ans.strip().lower() != "y":
            log.info("Отменено.")
            conn.close()
            return

    now = datetime.now(timezone.utc)

    rows_to_insert = []
    seen_pairs = set()
    skipped_no_data = 0
    skipped_too_long = 0
    multi_code_rows = 0

    log.info("Читаем строки…")
    for excel_row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        agsk_code   = row[1].value   # колонка B
        enstru_cell = row[5].value   # колонка F

        if not agsk_code or not enstru_cell:
            skipped_no_data += 1
            continue

        agsk_code   = str(agsk_code).strip().strip("'")
        enstru_cell = str(enstru_cell).strip().strip("'")

        if not agsk_code or not enstru_cell:
            skipped_no_data += 1
            continue

        # Одна ячейка F может содержать несколько кодов ЕНСТРУ через запятую.
        enstru_codes = split_enstru_codes(enstru_cell)
        if not enstru_codes:
            skipped_no_data += 1
            continue
        if len(enstru_codes) > 1:
            multi_code_rows += 1

        analyst_id  = analyst_for_row(excel_row_num)
        is_green    = (row_color_key(row[0]) == GREEN)
        is_approved = is_green
        approved_by = analyst_id if is_green else None
        approved_at = now        if is_green else None

        for enstru_code in enstru_codes:
            if len(enstru_code) > ENSTRU_MAX_LEN:
                log.warning(
                    "Строка %d: код ЕНСТРУ длиннее %d символов, пропущен: %r",
                    excel_row_num, ENSTRU_MAX_LEN, enstru_code,
                )
                skipped_too_long += 1
                continue

            pair = (agsk_code, enstru_code)
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)

            rows_to_insert.append((
                agsk_code,
                enstru_code,
                analyst_id,   # created_by
                is_approved,
                approved_by,
                now,          # created_at
                approved_at,
                True,         # is_active
            ))

        if excel_row_num % 20000 == 0:
            log.info("  обработано строк: %d  / собрано записей: %d", excel_row_num, len(rows_to_insert))

    log.info(
        "Всего записей для вставки: %d  (строк с несколькими кодами: %d, "
        "пропущено пустых: %d, пропущено слишком длинных: %d)",
        len(rows_to_insert), multi_code_rows, skipped_no_data, skipped_too_long,
    )

    BATCH = 5000
    inserted_total = 0
    for start in range(0, len(rows_to_insert), BATCH):
        batch = rows_to_insert[start : start + BATCH]
        execute_values(
            cur,
            """
            INSERT INTO agsk_enstru_matches
                (agsk_code, enstru_code, created_by, is_approved, approved_by,
                 created_at, approved_at, is_active)
            VALUES %s
            ON CONFLICT DO NOTHING
            """,
            batch,
        )
        conn.commit()
        inserted_total += len(batch)
        log.info("  вставлено: %d / %d", inserted_total, len(rows_to_insert))

    cur.execute("SELECT COUNT(*) FROM agsk_enstru_matches")
    total_in_db = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM agsk_enstru_matches WHERE is_approved = TRUE")
    approved_in_db = cur.fetchone()[0]

    log.info("─── Готово ───────────────────────────────────────────────────")
    log.info("Всего в таблице:     %d", total_in_db)
    log.info("  из них утверждено: %d (зелёные)", approved_in_db)
    log.info("  ожидает:           %d (оранж/фиол/голубые)", total_in_db - approved_in_db)

    cur.close()
    conn.close()


if __name__ == "__main__":
    main()
