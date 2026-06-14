"""Парсинг входящих ПСД-документов:
- `parse_psd_file` — XML (.kenml/.xml/.zip)
- `parse_smeta_file` — Excel-смета (.xlsx)

Используется как миксин в PsdAnalystService. После парсинга для XML-файлов
вызывается self._run_auto_matching_for_document (из PsdSearchMixin).
"""
import zipfile
from typing import Dict

from sqlalchemy.orm import Session

from ..models.models import Enstru, ExternalDocument, PsdDocumentItem


class PsdParserMixin:
    def parse_psd_file(self, db: Session, doc_id: int, file_path: str):
        import xml.etree.ElementTree as ET

        KEYS = {
            'type': ['Тип', 'Type', 'Kind', 'Class'],
            'code_snb': ['КодСНБ', 'CodeSNB', 'Cipher', 'Code'],
            'name': ['Наименование', 'Name', 'Description'],
            'unit': ['Измеритель', 'Unit', 'MeasureUnit', 'UOM', 'Measure'],
            'vol': ['Объем', 'Quantity', 'Volume', 'Amount', 'Count'],
        }

        def _strip_ns(tree):
            for el in tree.iter():
                if '}' in el.tag:
                    el.tag = el.tag.split('}', 1)[1]
            return tree

        def _f(val):
            if not val: return 0.0
            try:
                return float(str(val).replace(',', '.').replace('\xa0', '').strip())
            except:
                return 0.0

        def _v(node, key):
            for k in KEYS.get(key, []):
                if node.get(k) is not None: return node.get(k)
            return None

        def _classify(xml_type, code_snb, unit_raw, is_resource=False, is_k2=False):
            unit = str(unit_raw).lower().strip()
            code = str(code_snb).strip()
            xt = str(xml_type).lower()

            if any(x in unit for x in ['маш.-ч', 'маш.ч', 'маш-ч', 'маш/ч', 'mach-h']): return 'SERVICES'
            if any(x in unit for x in ['чел.-ч', 'чел.ч', 'чел-ч', 'чел/ч', 'man-h']): return 'WORKS'
            if any(x in unit for x in ['т·км', 'т•км', 'ткм', 'т/км']): return 'SERVICES'
            if any(x in unit for x in ['м3 подстилающего', 'м3 основания', 'м2 поверхности', 'м2 полотна', 'переезд']): return 'WORKS'

            if code.startswith('556') or code.startswith('557'): return 'GOODS'

            digits = code.replace('-', '')
            if digits[:4].isdigit():
                prefix = int(digits[:4])
                if 1000 <= prefix <= 1999: return 'WORKS'
                if 6000 <= prefix <= 6999: return 'WORKS'
                if 3000 <= prefix <= 4999: return 'SERVICES'

            if is_k2:
                if xt in ['2', '5', '5.1']: return 'GOODS'
                if xt in ['3', '4', '6']: return 'SERVICES'
                if xt in ['1', '1.1', '1.2', '0']: return 'WORKS'
            else:
                if not is_resource:
                    if xt in ['1', '2']: return 'GOODS'
                else:
                    if xt == '2': return 'GOODS'
                    if xt == '3': return 'SERVICES'
                if xt == '6': return 'SERVICES'

            if xt in ['material', 'материал', 'equipment', 'оборудование']: return 'GOODS'
            if xt in ['machine', 'mechanism', 'механизм']: return 'SERVICES'

            if not code:
                return 'OTHER'
            return 'WORKS'

        def _cost(node, is_k2=False):
            price, total = 0.0, 0.0
            if is_k2:
                cn = node.find('Cost')
                if cn is not None:
                    price = _f(cn.get('Est_Price') or cn.get('Summary'))
                    vp = cn.find('Volume_Price')
                    total = _f(vp.get('Summary')) if vp is not None else _f(cn.get('Summary'))
            else:
                cn = node.find('СТОИМОСТЬ')
                if cn is not None:
                    tn = cn.find('ВСЕГО')
                    un = cn.find('ЕДИНИЦА')
                    if tn is not None:
                        total = _f(tn.get('Всего') or tn.get('ПЗ') or tn.get('ОТП'))
                    if un is not None:
                        price = _f(un.get('Всего') or un.get('ПЗ') or un.get('Цена'))
                if total == 0:
                    fc = node.find('Cost')
                    if fc is not None: total = _f(fc.get('total') or fc.get('Summary'))
                if price == 0: price = _f(node.get('Цена'))
            return price, total

        all_rows = []

        def _process_xml(xml_text):
            try:
                root = _strip_ns(ET.ElementTree(ET.fromstring(xml_text))).getroot()
                is_k2 = root.tag in ['LS', 'OS', 'SSR']

                official_total = 0.0
                if is_k2:
                    itn = root.find('.//Summary_LS') or root.find('.//Summary_OS') or root.find('.//SummaryDoc')
                    if itn is not None: official_total = _f(itn.get('Summary'))
                else:
                    itn = root.find('ИТОГДОК')
                    if itn is not None:
                        official_total = _f(itn.get('Всего') or itn.get('ЗатратыПодрядчика'))
                    if official_total == 0:
                        official_total = _f(root.get('Всего') or 0)

                calculated_total = 0.0
                positions = list(dict.fromkeys(
                    root.findall('.//ПОЗИЦИЯ') +
                    root.findall('.//Position') +
                    root.findall('.//Item')
                ))

                for pos in positions:
                    p_type = _v(pos, 'type') or '0'
                    p_code = _v(pos, 'code_snb') or ''
                    p_unit = _v(pos, 'unit') or ''
                    p_name = _v(pos, 'name') or ''
                    p_vol = _f(_v(pos, 'vol'))
                    item_type = _classify(p_type, p_code, p_unit, is_resource=False, is_k2=is_k2)

                    p_price, p_total_xml = _cost(pos, is_k2)
                    if p_price > 0 and p_vol > 0:
                        p_total = p_price * p_vol
                    else:
                        p_total = p_total_xml
                        if p_vol > 0 and p_total > 0: p_price = p_total / p_vol

                    extracted = 0.0
                    if item_type == 'WORKS':
                        for res in (pos.findall('РЕСУРС') + pos.findall('Resource') + pos.findall('Subitem')):
                            r_type = _classify(
                                _v(res, 'type'), _v(res, 'code_snb') or '', _v(res, 'unit') or '',
                                is_resource=True, is_k2=is_k2
                            )
                            if r_type == 'WORKS': continue
                            r_vol = _f(_v(res, 'vol'))
                            r_price, r_total_xml = _cost(res, is_k2)
                            if r_price > 0 and r_vol > 0:
                                r_total = r_price * r_vol
                            else:
                                r_total = r_total_xml
                                if r_vol > 0 and r_total > 0: r_price = r_total / r_vol
                            extracted += r_total
                            all_rows.append({
                                'name': _v(res, 'name') or '', 'code': _v(res, 'code_snb') or '',
                                'unit': _v(res, 'unit') or '', 'vol': r_vol,
                                'price': r_price, 'total': r_total, 'item_type': r_type,
                            })
                            calculated_total += r_total

                    net_total = (p_total - extracted) if item_type == 'WORKS' else p_total
                    if item_type == 'WORKS' and net_total < 0 and abs(net_total) < 5.0:
                        net_total = 0.0
                    net_price = net_total / p_vol if p_vol > 0 else p_price

                    all_rows.append({
                        'name': p_name, 'code': p_code, 'unit': p_unit,
                        'vol': p_vol, 'price': net_price, 'total': net_total, 'item_type': item_type,
                    })
                    calculated_total += net_total

                delta = official_total - calculated_total
                if abs(delta) > 1.0:
                    all_rows.append({
                        'name': 'Разница (НДС, Округления, Лимиты)',
                        'code': 'БАЛАНС', 'unit': 'компл',
                        'vol': 1.0, 'price': delta, 'total': delta, 'item_type': 'BALANCE',
                    })
            except Exception:
                pass

        if file_path.lower().endswith('.zip'):
            with zipfile.ZipFile(file_path, 'r') as z:
                for name in z.namelist():
                    if not (name.lower().endswith('.kenml') or name.lower().endswith('.xml')):
                        continue
                    try:
                        content = z.read(name)
                        for enc in ['utf-8', 'cp1251']:
                            try:
                                txt = content.decode(enc)
                                if '<' in txt[:300]:
                                    _process_xml(txt)
                                    break
                            except:
                                pass
                    except:
                        pass
        else:
            try:
                with open(file_path, 'rb') as f:
                    content = f.read()
                for enc in ['utf-8', 'cp1251']:
                    try:
                        txt = content.decode(enc)
                        if '<' in txt[:300]:
                            _process_xml(txt)
                            break
                    except:
                        pass
            except:
                pass

        db.query(PsdDocumentItem).filter(PsdDocumentItem.document_id == doc_id).delete()

        from .psd_analyzer.analyzer import clean_product_name

        grouped = {}
        for row in all_rows:
            itype = row['item_type']
            code = str(row['code']).strip()
            name = str(row['name']).strip()
            if itype == 'BALANCE':
                key = f"BALANCE_{name}"
                grouped[key] = row
                continue
            key = (itype, code, name)
            if key not in grouped:
                grouped[key] = {**row, 'total': 0.0, 'vol': 0.0}
            grouped[key]['total'] += float(row['total'])
            grouped[key]['vol'] += float(row['vol'])

        final_rows = []
        for row in grouped.values():
            vol = row['vol']
            total = row['total']
            row['price'] = total / vol if vol > 0 else 0.0
            final_rows.append(row)

        for idx, row in enumerate(final_rows, 1):
            name = str(row['name']).strip() or f'Позиция {idx}'
            code = str(row['code']).strip()
            itype = row['item_type']
            db.add(PsdDocumentItem(
                document_id=doc_id,
                position_number=str(idx),
                name=name,
                code_sn=code if (code and code != 'БАЛАНС') else ('Прайс-лист' if itype == 'GOODS' else None),
                unit=str(row['unit']),
                volume=row['vol'],
                price=float(row['price']),
                total_amount=row['total'],
                clean_name=clean_product_name(name),
                is_product=(itype == 'GOODS'),
                item_type=itype,
                match_type='none',
            ))

        db.commit()

        doc = db.query(ExternalDocument).filter(ExternalDocument.id == doc_id).first()
        if doc and doc.status != 'ASSIGNED_TO_ANALYST':
            doc.status = 'PARSED'
            db.commit()

        self._run_auto_matching_for_document(db, doc_id)

    def parse_smeta_file(self, db: Session, doc_id: int, file_path: str):
        import openpyxl

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb["Позиции для загрузки"] if "Позиции для загрузки" in wb.sheetnames else wb.active
        db.query(PsdDocumentItem).filter(PsdDocumentItem.document_id == doc_id).delete()

        # ── Шаг 1: собираем все строки в память + копим уникальные ЕНСТРУ-коды
        rows_to_process: list = []
        enstru_codes_seen: set = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            row_data = list(row) + [None] * max(0, 17 - len(row))
            enstru_code = str(row_data[1]).strip() if row_data[1] else None
            if enstru_code:
                enstru_codes_seen.add(enstru_code)
            rows_to_process.append(row_data)

        # ── Шаг 2: один батч-запрос в справочник ЕНСТРУ — name_rus по коду
        enstru_name_map: Dict[str, str] = {}
        if enstru_codes_seen:
            for e in db.query(Enstru.code, Enstru.name_rus).filter(Enstru.code.in_(enstru_codes_seen)).all():
                if e.name_rus:
                    enstru_name_map[e.code] = e.name_rus

        # ── Шаг 2.5: какие ЕНСТРУ-коды из сметы реально присутствуют у активных
        # поставщиков с валидным ДВС в реестре КТП. Только для них имеет смысл
        # «подсказка» — иначе аналитику нечего будет выбирать (UI-поиск
        # фильтрует поставщиков без ДВС, и тогда ничего не найдётся).
        enstru_in_ktp = self._enstru_codes_with_active_supplier(db, enstru_codes_seen)

        # ── Шаг 3: создаём позиции
        position_idx = 1
        for row_data in rows_to_process:
            enstru_code = str(row_data[1]).strip() if row_data[1] else None
            col_c = str(row_data[2]).strip() if row_data[2] else None  # краткое имя в смете
            col_d = str(row_data[3]).strip() if row_data[3] else None  # наименование позиции

            # Приоритет наименования: справочник ЕНСТРУ → колонка C → колонка D → автогенерация
            name = None
            if enstru_code and enstru_code in enstru_name_map:
                name = enstru_name_map[enstru_code]
            if not name:
                name = col_c or col_d or f"Позиция {position_idx}"

            unit = str(row_data[5]).strip() if row_data[5] else ""
            agsk_code = str(row_data[13]).strip() if row_data[13] else None
            if agsk_code and agsk_code.lower() == "прайс-лист":
                agsk_code = None
            try:
                volume = float(row_data[6]) if row_data[6] is not None else 0.0
            except (ValueError, TypeError):
                volume = 0.0
            try:
                price = float(row_data[7]) if row_data[7] is not None else 0.0
            except (ValueError, TypeError):
                price = 0.0
            try:
                total_amount = float(row_data[8]) if row_data[8] is not None else (volume * price)
            except (ValueError, TypeError):
                total_amount = volume * price

            if not enstru_code and not name:
                continue

            # ЕНСТРУ-код из сметы — это ПОДСКАЗКА, а не авто-сопоставление.
            # Аналитик всё равно должен выбрать конкретного поставщика из реестра КТП
            # (как с библиотечной подсказкой). Авто-статус ставит только AGSK→КТП matcher.
            #
            # Подсказку («suggested») ставим ТОЛЬКО если этот ЕНСТРУ реально есть
            # у активных поставщиков в реестре КТП — иначе выбирать всё равно не из
            # кого, лучше показать «не указано» (match_type='none').
            has_supplier_in_ktp = bool(enstru_code and enstru_code in enstru_in_ktp)
            item = PsdDocumentItem(
                document_id=doc_id, position_number=str(position_idx), name=name,
                code_sn=agsk_code, unit=unit, volume=volume, price=price,
                total_amount=total_amount,
                clean_name=name.split('/')[0].strip() if '/' in name else name,
                is_product=True,
                item_type='GOODS',
                enstru_code=enstru_code,
                enstru_name=enstru_name_map.get(enstru_code) if enstru_code else None,
                match_type="suggested" if has_supplier_in_ktp else "none",
                match_score=100.0 if has_supplier_in_ktp else None,
                match_reason=(
                    "ЕНСТРУ из сметы — выберите поставщика из реестра КТП"
                    if has_supplier_in_ktp
                    else ("ЕНСТРУ из сметы, но в реестре КТП нет активных поставщиков" if enstru_code else None)
                ),
            )
            db.add(item)
            position_idx += 1

        db.commit()
        doc = db.query(ExternalDocument).filter(ExternalDocument.id == doc_id).first()
        if doc and doc.status != "ASSIGNED_TO_ANALYST":
            doc.status = "PARSED"
            db.commit()

        # Запускаем авто-матчинг АГСК → КТП — для позиций сметы с АГСК-кодом,
        # который напрямую есть в реестре КТП, match_type станет auto/auto_ktp.
        # Позиции с ЕНСТРУ из сметы, но без АГСК в КТП — останутся 'suggested'
        # (логика в _run_auto_matching_for_document).
        self._run_auto_matching_for_document(db, doc_id)
