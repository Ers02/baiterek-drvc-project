import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname
from sqlalchemy.orm import Session
from ...models import models


def generate_import_template(db: Session) -> bytes:
    """Генерирует Excel-шаблон с отдельными листами для справочников и именованными диапазонами."""
    wb = openpyxl.Workbook()
    
    default_sheet = wb.active
    wb.remove(default_sheet)

    def create_ref_sheet(sheet_name, data_list, range_name):
        ws = wb.create_sheet(sheet_name)
        ws.sheet_state = 'hidden'
        
        for idx, val in enumerate(data_list, start=1):
            ws.cell(row=idx, column=1, value=val)
        
        if data_list:
            quoted_name = quote_sheetname(sheet_name)
            formula = f"{quoted_name}!$A$1:$A${len(data_list)}"
            d_name = DefinedName(range_name, attr_text=formula)
            wb.defined_names.add(d_name)

    mkeis = db.query(models.Mkei.code, models.Mkei.name_ru).all()
    mkei_data = [f"{code} - {name}" for code, name in mkeis]
    create_ref_sheet("Ref_MKEI", mkei_data, "List_MKEI")

    cost_items = db.query(models.Cost_Item.id, models.Cost_Item.name_ru).all()
    cost_data = [f"{id} - {name}" for id, name in cost_items]
    create_ref_sheet("Ref_Cost", cost_data, "List_Cost")

    sources = db.query(models.Source_Funding.id, models.Source_Funding.name_ru).all()
    source_data = [f"{id} - {name}" for id, name in sources]
    create_ref_sheet("Ref_Source", source_data, "List_Source")

    katos = db.query(models.Kato.code, models.Kato.name_ru).all()
    kato_data = [f"{code} - {name}" for code, name in katos]
    create_ref_sheet("Ref_KATO", kato_data, "List_KATO")

    ws_data = wb.create_sheet("Позиции для загрузки", 0)
    
    headers = [
        "№",
        "Код по ЕНС ТРУ",
        "Наименование закупаемых товаров, работ и услуг",
        "Дополнительная характеристика",
        "Единица измерения(МКЕИ) (для товаров)",
        "Количество, объем",
        "Цена за единицу, тенге без НДС",
        "Сумма планируемая для закупок ТРУ без НДС, тенге",
        "Место закупки (КАТО)",
        "Место поставки (КАТО)",
        "Статья затрат",
        "Источник финансирования",
        "Код АГСК (для СМР)\nИз справочника АГСК-3",
        "Доля внутристрановой ценности (%)",
        "Обоснование если доля внутристрановой ценности ниже 100%",
        "Сумма ВЦ (тенге)"
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws_data.append(headers)
    ws_data.row_dimensions[1].height = 45

    for cell in ws_data[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    widths = {'A': 5, 'B': 20, 'C': 20, 'D': 20, 'E': 20, 'F': 20, 'G': 15, 'H': 15, 
              'I': 20, 'J': 21, 'K': 21, 'L': 21, 'M': 21, 'N': 22, 'O': 17, 'P': 30, 'Q': 20}
    for col, width in widths.items():
        ws_data.column_dimensions[col].width = width

    data_rows_count = 2000 

    def add_dv(formula_name, col_letter):
        dv = DataValidation(type="list", formula1=f"={formula_name}", allow_blank=True)
        dv.error = 'Выберите значение из списка'
        dv.errorTitle = 'Неверное значение'
        ws_data.add_data_validation(dv)
        dv.add(f'{col_letter}2:{col_letter}{data_rows_count}')

    if mkei_data: add_dv("List_MKEI", "F")
    if kato_data: 
        add_dv("List_KATO", "J")
        add_dv("List_KATO", "K")
    if cost_data: add_dv("List_Cost", "L")
    if source_data: add_dv("List_Source", "M")

    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    return virtual_workbook.getvalue()


def generate_error_report(errors: list) -> bytes:
    """Генерирует Excel-файл с отчетом об ошибках."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ошибки импорта"
    
    headers = ["Номер строки", "Описание ошибки"]
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    for err in errors:
        ws.append([err['row'], err['message']])
        
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 100
    
    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    return virtual_workbook.getvalue()
