import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy import func
from fastapi import HTTPException

from ...models import models
from ...utils.helpers import is_smr


def export_plan_to_excel(db: Session, plan_id: int, version_id: int = None) -> bytes:
    """
    Экспортирует версию плана в Excel.
    """
    query = db.query(models.ProcurementPlanVersion).filter(models.ProcurementPlanVersion.plan_id == plan_id)
    
    if version_id:
        query = query.filter(models.ProcurementPlanVersion.id == version_id)
    else:
        query = query.filter(models.ProcurementPlanVersion.is_active == True)

    version = query.first()

    if not version:
        raise HTTPException(status_code=404, detail="Версия сметы не найдена")

    version_with_items = db.query(models.ProcurementPlanVersion).options(
        selectinload(models.ProcurementPlanVersion.items).options(
            joinedload(models.PlanItemVersion.enstru),
            joinedload(models.PlanItemVersion.unit),
            joinedload(models.PlanItemVersion.expense_item),
            joinedload(models.PlanItemVersion.funding_source),
            joinedload(models.PlanItemVersion.agsk),
            joinedload(models.PlanItemVersion.kato_purchase),
            joinedload(models.PlanItemVersion.kato_delivery),
            joinedload(models.PlanItemVersion.source_version),
            joinedload(models.PlanItemVersion.root_item).joinedload(models.PlanItemVersion.version)
        ),
        joinedload(models.ProcurementPlanVersion.plan).joinedload(models.ProcurementPlan.creator)
    ).filter(models.ProcurementPlanVersion.id == version.id).one()

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    sub_header_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def format_item_number(idx, item):
        number = f"{idx}"
        if item.revision_number > 0:
            number += f"-{item.revision_number}"

        type_suffix = ""
        if item.need_type == models.NeedType.GOODS: type_suffix = " Т"
        elif item.need_type == models.NeedType.WORKS: type_suffix = " Р"
        elif item.need_type == models.NeedType.SERVICES: type_suffix = " У"

        return f"{number}{type_suffix}"

    grouped_items = {
        models.NeedType.GOODS: [],
        models.NeedType.WORKS: [],
        models.NeedType.SERVICES: []
    }

    for item in version_with_items.items:
        if not item.is_deleted:
            grouped_items[item.need_type].append(item)

    for key in grouped_items:
        grouped_items[key].sort(key=lambda x: x.item_number)

    ws = wb.active
    ws.title = "Смета"

    ws.merge_cells('A1:Q1')
    ws['A1'] = "СМЕТА ЗАКУПОК"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:Q2')
    ws['A2'] = f"Наименование проекта: {version_with_items.plan.plan_name}"
    ws['A2'].font = Font(bold=True, size=12)

    ws.merge_cells('A3:Q3')
    ws['A3'] = f"Год: {version_with_items.plan.year}"
    ws['A3'].font = Font(bold=True, size=12)

    client_name = version_with_items.plan.creator.org_name if version_with_items.plan.creator and version_with_items.plan.creator.org_name else "Не указано"
    ws.merge_cells('A4:Q4')
    ws['A4'] = f"Наименование клиента: {client_name}"
    ws['A4'].font = Font(bold=True, size=12)

    current_row = 6

    columns = [
        "№",
        "Код по ЕНС ТРУ",
        "Наименование закупаемых товаров услуг работ",
        "Краткая характеристика",
        "Дополнительная характеристика",
        "Единица измерения(МКЕИ)",
        "Количество, объём",
        "Цена за единицу тенге(без НДС)",
        "Сумма планируемая для закупок ТРУ",
        "Место закупки(КАТО)",
        "Место поставки(КАТО)",
        "Статья затрат",
        "Источник финансирования",
        "КОД АГСК для смр",
        "КТП",
        "ВЦ %",
        "Сумма ВЦ тенге без НДС"
    ]

    def create_table_header(ws, row_idx, cols):
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[row_idx].height = 45
        return row_idx + 1

    def fill_section(title, items, start_row, is_first_section=False):
        if not items: return start_row
        
        if is_first_section:
            start_row = create_table_header(ws, start_row, columns)
        
        ws.merge_cells(f'A{start_row}:Q{start_row}')
        ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
        ws.cell(row=start_row, column=1).fill = sub_header_fill
        start_row += 1
        
        section_total = Decimal('0.00')
        section_vc_amount = Decimal('0.00')
        
        # Получаем min_dvc для всех товаров в этой секции
        # Теперь используем готовое поле min_dvc_percent из PlanItemVersion
        min_dvc_map = {}
        if items and items[0].need_type == models.NeedType.GOODS:
            for item in items:
                if item.min_dvc_percent is not None:
                    min_dvc_map[item.trucode] = item.min_dvc_percent

        for idx, item in enumerate(items, 1):
            agsk_value = ""
            if is_smr(item.expense_item_id):
                if item.agsk_code:
                    agsk_value = item.agsk_code
                else:
                    agsk_value = "Прайс-лист"
            elif item.agsk_code:
                agsk_value = item.agsk_code
            
            unit_display = item.unit.name_ru if item.unit else (item.original_unit_name or "")

            # Определяем ВЦ для экспорта
            if item.need_type == models.NeedType.GOODS:
                dvc_percent = min_dvc_map.get(item.trucode, Decimal(0))
            else:
                dvc_percent = item.resident_share

            # Используем готовое поле vc_amount из модели
            vc_amount = item.vc_amount

            row_data = [
                format_item_number(idx, item),
                item.trucode,
                item.enstru.name_rus if item.enstru else "",
                item.enstru.detail_rus if item.enstru else "",
                item.additional_specs,
                unit_display,
                item.quantity,
                item.price_per_unit,
                item.total_amount,
                item.kato_purchase.name_ru if item.kato_purchase else "",
                item.kato_delivery.name_ru if item.kato_delivery else "",
                item.expense_item.name_ru if item.expense_item else "",
                item.funding_source.name_ru if item.funding_source else "",
                agsk_value,
                "Да" if item.is_ktp else "Нет",
                f"{dvc_percent}",
                vc_amount
            ]
            
            section_total += item.total_amount
            section_vc_amount += vc_amount
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=val)
                cell.border = border
                if col_idx in [7, 8, 9, 17]:
                    cell.number_format = '#,##0.00'
            
            start_row += 1
            
        ws.merge_cells(f'A{start_row}:H{start_row}')
        ws.cell(row=start_row, column=1, value=f"Итого по {title.lower()}:").font = Font(bold=True)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='right')
        ws.cell(row=start_row, column=9, value=section_total).font = Font(bold=True)
        ws.cell(row=start_row, column=9).number_format = '#,##0.00'
        
        section_vc_mean = (section_vc_amount / section_total * 100) if section_total > 0 else Decimal('0.00')
        
        ws.cell(row=start_row, column=16, value=f"{section_vc_mean.quantize(Decimal('0.00'))}%").font = Font(bold=True)
        ws.cell(row=start_row, column=17, value=section_vc_amount).font = Font(bold=True)
        ws.cell(row=start_row, column=17).number_format = '#,##0.00'
        
        return start_row + 2

    current_row = fill_section("1. Товары", grouped_items[models.NeedType.GOODS], current_row, is_first_section=True)
    current_row = fill_section("2. Работы", grouped_items[models.NeedType.WORKS], current_row)
    current_row = fill_section("3. Услуги", grouped_items[models.NeedType.SERVICES], current_row)

    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws.cell(row=current_row, column=1, value="Всего:").font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right')
    ws.cell(row=current_row, column=9, value=version_with_items.total_amount).font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=9).number_format = '#,##0.00'
    
    total_vc_mean = (version_with_items.vc_amount / version_with_items.total_amount * 100) if version_with_items.total_amount > 0 else Decimal('0.00')

    current_row += 1
    ws.cell(row=current_row, column=9, value="Средний % ВЦ:").font = Font(bold=True)
    ws.cell(row=current_row, column=10, value=f"{total_vc_mean.quantize(Decimal('0.00'))}%").font = Font(bold=True)

    current_row += 1
    ws.cell(row=current_row, column=9, value="Общая сумма ВЦ:").font = Font(bold=True)
    ws.cell(row=current_row, column=10, value=version_with_items.vc_amount).font = Font(bold=True)
    ws.cell(row=current_row, column=10).number_format = '#,##0.00'
    
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)


    ws_ktp = wb.create_sheet("КТП")

    ktp_columns = [
        "№",
        "Код по ЕНС ТРУ",
        "Наименование закупаемых товаров услуг работ",
        "Краткая характеристика",
        "Дополнительная характеристика",
        "Единица измерения(МКЕИ)",
        "Количество, объём",
        "Цена за единицу тенге(без НДС)",
        "Сумма планируемая для закупок ТРУ",
        "Место закупки(КАТО)",
        "Место поставки(КАТО)",
        "Статья затрат",
        "Источник финансирования",
        "Код АГСК-3 для СМР",
        "КТП",
        "Сумма ВЦ тенге без НДС",
        "БИН производителя",
        "Наименования производителя",
        "Адрес/ контакты",
        "ВЦ% по этому производителю",
        "Сумма ВЦ тенге без НДС (по производителю)"
    ]

    ktp_row = 1
    ktp_row = create_table_header(ws_ktp, ktp_row, ktp_columns)

    for t in [models.NeedType.GOODS, models.NeedType.WORKS, models.NeedType.SERVICES]:
        items = grouped_items[t]
        for idx, item in enumerate(items, 1):
            suppliers = db.query(models.Reestr_KTP).filter(
                models.Reestr_KTP.enstru_codes.contains([item.trucode])
            ).all()

            if suppliers:
                for supplier in suppliers:
                    # dvc_percent это текст, приводим к float
                    try:
                        supplier_dvc = float(supplier.dvc_percent.replace(',', '.')) if supplier.dvc_percent else 0.0
                    except ValueError:
                        supplier_dvc = 0.0
                    
                    # Показываем только поставщиков с ВЦ больше 0
                    if supplier_dvc <= 0:
                        continue
                    
                    supplier_vc_amount = item.total_amount * (Decimal(str(supplier_dvc)) / Decimal('100.00'))

                    agsk_value = ""
                    if is_smr(item.expense_item_id):
                        if item.agsk_code:
                            agsk_value = item.agsk_code
                        else:
                            agsk_value = "Прайс-лист"
                    elif item.agsk_code:
                        agsk_value = item.agsk_code

                    row_data = [
                        format_item_number(idx, item),
                        item.trucode,
                        item.enstru.name_rus if item.enstru else "",
                        item.enstru.detail_rus if item.enstru else "",
                        item.additional_specs,
                        item.unit.name_ru if item.unit else "",
                        item.quantity,
                        item.price_per_unit,
                        item.total_amount,
                        item.kato_purchase.name_ru if item.kato_purchase else "",
                        item.kato_delivery.name_ru if item.kato_delivery else "",
                        item.expense_item.name_ru if item.expense_item else "",
                        item.funding_source.name_ru if item.funding_source else "",
                        agsk_value,
                        "Да" if item.is_ktp else "Нет",
                        # Используем готовое поле vc_amount из модели
                        item.vc_amount,

                        supplier.bin_iin,
                        supplier.company_name,
                        f"{supplier.production_address or ''} {supplier.phone or ''} {supplier.email or ''}",
                        f"{supplier_dvc}",
                        supplier_vc_amount
                    ]

                    for col_idx, val in enumerate(row_data, 1):
                        cell = ws_ktp.cell(row=ktp_row, column=col_idx, value=val)
                        cell.border = border
                        if col_idx in [7, 8, 9, 16, 21]:
                            cell.number_format = '#,##0.00'

                    ktp_row += 1

    for i, col in enumerate(ws_ktp.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_ktp.column_dimensions[column_letter].width = min(adjusted_width, 50)

    ws_nr = wb.create_sheet("Услуги, Работы ВЦ меньше 100%")

    nr_columns = [
        "№",
        "Код по ЕНС ТРУ",
        "Наименование закупаемых товаров услуг работ",
        "Краткая характеристика",
        "Дополнительная характеристика",
        "Единица измерения(МКЕИ)",
        "Количество, объём",
        "Цена за единицу тенге без НДС",
        "Сумма планируемая для закупок ТРУ",
        "Место закупки(КАТО)",
        "Место поставки(КАТО)",
        "Статья затрат",
        "Источник финансирования",
        "Код АГСК-3 для СМР",
        "Доля внутристрановой ценности (%)",
        "Обоснование если доля внутристрановой ценности меньше 100%"
    ]

    nr_row = 1
    nr_row = create_table_header(ws_nr, nr_row, nr_columns)

    for t in [models.NeedType.WORKS, models.NeedType.SERVICES]:
        items = grouped_items[t]
        for idx, item in enumerate(items, 1):
            if item.resident_share < 100:
                agsk_value = ""
                if is_smr(item.expense_item_id):
                    if item.agsk_code:
                        agsk_value = item.agsk_code
                    else:
                        agsk_value = "Прайс-лист"
                elif item.agsk_code:
                    agsk_value = item.agsk_code

                row_data = [
                    format_item_number(idx, item),
                    item.trucode,
                    item.enstru.name_rus if item.enstru else "",
                    item.enstru.detail_rus if item.enstru else "",
                    item.additional_specs,
                    item.unit.name_ru if item.unit else "",
                    item.quantity,
                    item.price_per_unit,
                    item.total_amount,
                    item.kato_purchase.name_ru if item.kato_purchase else "",
                    item.kato_delivery.name_ru if item.kato_delivery else "",
                    item.expense_item.name_ru if item.expense_item else "",
                    item.funding_source.name_ru if item.funding_source else "",
                    agsk_value,
                    f"{item.resident_share}",
                    item.non_resident_reason
                ]

                for col_idx, val in enumerate(row_data, 1):
                    cell = ws_nr.cell(row=nr_row, column=col_idx, value=val)
                    cell.border = border
                    if col_idx in [7, 8, 9]:
                        cell.number_format = '#,##0.00'

                nr_row += 1

    for i, col in enumerate(ws_nr.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_nr.column_dimensions[column_letter].width = min(adjusted_width, 50)

    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    return virtual_workbook.getvalue()
