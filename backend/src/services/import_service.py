import io
import os
import uuid
import time
from decimal import Decimal
from sqlalchemy.orm import Session
from fastapi import UploadFile, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse

from ..models import models
from ..services import plan_service
from .exporters.excel_generator import generate_import_template, generate_error_report
from .importers.excel_importer import import_items_from_excel
from .importers.kenml_parser import parse_kenml_file
from ..core.config import settings
from ..core.logger import logger

__all__ = ['generate_import_template', 'process_import_file', 'process_kenml_import']


def save_error_report(errors: list, filename: str):
    """Сохраняет отчет об ошибках на диск."""
    if not os.path.exists(settings.REPORT_DIR):
        os.makedirs(settings.REPORT_DIR)

    content = generate_error_report(errors)
    path = os.path.join(settings.REPORT_DIR, filename)
    with open(path, "wb") as f:
        f.write(content)
    return path


def delete_file(path: str):
    """Удаляет файл с диска."""
    try:
        if os.path.exists(path):
            os.remove(path)
            logger.info(f"Deleted file: {path}")
    except Exception as e:
        logger.error(f"Error deleting file {path}: {e}")


def process_import_file(
    db: Session, 
    plan_id: int, 
    file: UploadFile, 
    user: models.User, 
    background_tasks: BackgroundTasks
):
    """
    Синхронный импорт Excel.
    Возвращает JSON при успехе или Excel-файл при ошибках валидации.
    """
    start_time = time.time()
    logger.info(f"Starting import for plan {plan_id}")

    active_version = plan_service._get_active_version(db, plan_id)
    if not active_version:
        raise HTTPException(status_code=404, detail="Активная версия плана не найдена")
    if active_version.status != models.PlanStatus.DRAFT:
        raise HTTPException(status_code=403, detail="Импорт возможен только в черновик")
    if active_version.plan.created_by != user.id:
        raise HTTPException(status_code=403, detail="Нет прав на редактирование этого плана")

    try:
        contents = file.file.read()
    except Exception as e:
        logger.error(f"Error reading file upload: {e}")
        raise HTTPException(status_code=400, detail="Ошибка чтения файла")

    t1 = time.time()
    new_items, errors = import_items_from_excel(db, active_version.id, contents)
    logger.info(f"Excel parsing took {time.time() - t1:.2f}s. Items: {len(new_items)}, Errors: {len(errors)}")

    if errors:
        filename = f"import_errors_{uuid.uuid4()}.xlsx"
        path = save_error_report(errors, filename)
        
        # Планируем удаление после отправки
        background_tasks.add_task(delete_file, path)
        
        return FileResponse(
            path,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename="import_errors.xlsx"
        )

    if not new_items:
        raise HTTPException(status_code=400, detail="Файл пуст или не содержит корректных данных")

    try:
        t2 = time.time()
        db.add_all(new_items)
        db.flush()
        logger.info(f"DB flush took {time.time() - t2:.2f}s")
        
        t3 = time.time()
        for item in new_items:
            item.root_item_id = item.id
        logger.info(f"Setting root_item_id took {time.time() - t3:.2f}s")
        
        t4 = time.time()
        plan_service._recalculate_version_metrics(db, active_version.id)
        logger.info(f"Recalculate metrics took {time.time() - t4:.2f}s")
        
        logger.info(f"Total import time: {time.time() - start_time:.2f}s")
        
        return JSONResponse(content={"message": f"Успешно импортировано {len(new_items)} позиций"})
        
    except Exception as e:
        db.rollback()
        logger.error(f"Import DB error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Ошибка при сохранении данных: {str(e)}")


def clean_product_name(name: str) -> str:
    if not name: return ""
    return name.split('/')[0].strip()


def process_kenml_import(db: Session, file: UploadFile):
    """
    Парсит KENML/ZIP, агрегирует данные и генерирует Excel-шаблон.
    Синхронный режим. Fuzzy Search отключен для скорости.
    """
    try:
        all_data = parse_kenml_file(file)
    except Exception as e:
        logger.error(f"KENML parse error: {e}")
        raise HTTPException(status_code=400, detail=f"Ошибка обработки файла: {str(e)}")

    goods = {}
    works_services = {}
    
    for row in all_data:
        if row['Сумма'] <= 0.01: continue
        
        cat = row['Категория']
        name = row['Наименование']
        code = row['КодСНБ']
        unit = row['Ед. изм.']
        vol = row['Объем']
        total = row['Сумма']
        
        if cat == 'Товары':
            key = (cat, name, code, unit)
            if key not in goods:
                goods[key] = {'vol': 0.0, 'total': 0.0}
            goods[key]['vol'] += vol
            goods[key]['total'] += total
        else:
            key = (cat, name, code)
            if key not in works_services:
                works_services[key] = {'total': 0.0}
            works_services[key]['total'] += total

    final_rows = []
    
    for key, val in goods.items():
        cat, name, code, unit = key
        vol = val['vol']
        total = val['total']
        price = total / vol if vol else 0
        final_rows.append({'cat': cat, 'name': name, 'code_agsk': code, 'unit': unit, 'vol': vol, 'price': price, 'total': total})
        
    for key, val in works_services.items():
        cat, name, code = key
        total = val['total']
        final_rows.append({'cat': cat, 'name': name, 'code_agsk': code, 'unit': '', 'vol': 1.0, 'price': total, 'total': total})
        
    final_rows.sort(key=lambda x: x['total'], reverse=True)
    
    # Поиск по АГСК (Обновлено для JSONB)
    agsk_codes_to_find = set(r['code_agsk'] for r in final_rows if r['code_agsk'] and r['code_agsk'] != "БЕЗ_КОДА")
    agsk_to_enstru = {}
    
    if agsk_codes_to_find:
        # Fetch relevant records from KTP
        # Note: Filtering JSONB array containment efficiently without index or specific operator is hard in pure ORM for 'any of list'.
        # We will fetch records that have agsk3_codes and filter in python.
        ktp_records = db.query(models.Reestr_KTP.agsk3_codes, models.Reestr_KTP.enstru_codes)\
            .filter(models.Reestr_KTP.agsk3_codes.isnot(None)).all()
            
        for rec in ktp_records:
            if not rec.agsk3_codes or not rec.enstru_codes: continue
            
            # Map each AGSK code in the record to the first ENSTRU code
            first_enstru = rec.enstru_codes[0]
            for agsk in rec.agsk3_codes:
                if agsk in agsk_codes_to_find:
                    agsk_to_enstru[agsk] = first_enstru

    # Сбор данных для Excel (ЕНС ТРУ и ДВЦ)
    all_found_enstru_codes = set(agsk_to_enstru.values())
    enstru_names = {}
    enstru_dvc = {}
    
    if all_found_enstru_codes:
        # Get names from ENSTRU dict
        for code, name in db.query(models.Enstru.code, models.Enstru.name_rus).filter(models.Enstru.code.in_(all_found_enstru_codes)).all():
            enstru_names[code] = name
            
        # Get Max DVC from KTP (Updated logic for JSONB/Text)
        ktp_candidates = db.query(models.Reestr_KTP.enstru_codes, models.Reestr_KTP.dvc_percent)\
            .filter(models.Reestr_KTP.enstru_codes.isnot(None), models.Reestr_KTP.dvc_percent.isnot(None)).all()
            
        for cand in ktp_candidates:
            if not cand.enstru_codes: continue
            try:
                dvc = Decimal(str(cand.dvc_percent).replace(',', '.'))
            except:
                continue
                
            # Check overlap
            common = set(cand.enstru_codes).intersection(all_found_enstru_codes)
            for code in common:
                if code not in enstru_dvc or dvc > enstru_dvc[code]:
                    enstru_dvc[code] = dvc

    # Генерация Excel
    import openpyxl
    template_bytes = generate_import_template(db)
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    ws = wb["Позиции для загрузки"]
    
    smr_cost_item = db.query(models.Cost_Item).filter(models.Cost_Item.id == settings.SMR_COST_ITEM_ID).first()
    smr_str = f"{smr_cost_item.id} - {smr_cost_item.name_ru}" if smr_cost_item else ""

    for idx, row in enumerate(final_rows, start=1):
        excel_row = idx + 1
        
        enstru_code = agsk_to_enstru.get(row['code_agsk'], "")
        
        ws.cell(row=excel_row, column=1, value=idx)
        ws.cell(row=excel_row, column=2, value=enstru_code)
        ws.cell(row=excel_row, column=3, value=enstru_names.get(enstru_code, ""))
        ws.cell(row=excel_row, column=4, value=row['name'])
        ws.cell(row=excel_row, column=5, value=row['name'])
        ws.cell(row=excel_row, column=6, value=row['unit'])
        ws.cell(row=excel_row, column=7, value=row['vol'])
        ws.cell(row=excel_row, column=8, value=row['price'])
        ws.cell(row=excel_row, column=9, value=row['total'])
        
        if smr_str: ws.cell(row=excel_row, column=12, value=smr_str)
        
        ws.cell(row=excel_row, column=14, value=row['code_agsk'])
        
        dvc_percent = enstru_dvc.get(enstru_code)
        if dvc_percent is not None:
             ws.cell(row=excel_row, column=15, value=dvc_percent)
             vc_amount = row['total'] * (float(dvc_percent) / 100.0)
             ws.cell(row=excel_row, column=17, value=vc_amount)
        
    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    return virtual_workbook.getvalue()
