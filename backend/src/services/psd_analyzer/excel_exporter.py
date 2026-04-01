import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from ...core.config import settings

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')


def clean_text_for_excel(text, max_len=5000):
    if text is None:
        return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    if len(text) > max_len:
        return text[:max_len] + "... (обрезано)"
    return text


# Цвета для подсветки качества совпадения
COLOR_EXACT = "C6EFCE"      # Зелёный — точное совпадение (100%, по АГСК)
COLOR_HIGH = "FFEB9C"       # Жёлтый — высокое (85-99%)
COLOR_LOW = "FFC7CE"        # Красный — низкое или не найдено (<85%)
COLOR_SKIPPED = "D9D9D9"    # Серый — пропущено (не товар)
COLOR_HEADER = "4472C4"     # Синий заголовок


class ExcelExporter:
    def __init__(self, items: list):
        self.items = items

    def _get_row_color(self, item: dict) -> str:
        if item.get('skip_search'):
            return COLOR_SKIPPED
        if not item.get('found_enstru'):
            return COLOR_LOW
        sim = item.get('similarity') or 0
        if sim >= 100:
            return COLOR_EXACT
        if sim >= 85:
            return COLOR_HIGH
        return COLOR_LOW

    def generate(self, task_id: str) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Анализ ПСД"

        headers = [
            "№", "Код ЕНС ТРУ", "Наименование (Смета)",
            "Наименование (Реестр КТП)", "Ед. изм.",
            "Кол-во", "Цена", "Сумма", "Код АГСК",
            "Схожесть (%)", "Статус", "Поставщики (Реестр КТП)"
        ]

        # Заголовок
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
        header_alignment = Alignment(horizontal="center", wrap_text=True)

        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Ширины колонок
        col_widths = {
            'A': 5, 'B': 22, 'C': 45, 'D': 45,
            'E': 8, 'F': 10, 'G': 14, 'H': 16,
            'I': 18, 'J': 12, 'K': 14, 'L': 60
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Данные
        for idx, item in enumerate(self.items, start=1):
            price = item['total'] / item['vol'] if item['vol'] else 0

            # Статус для читаемости
            if item.get('skip_search'):
                status = "Пропущено"
            elif not item.get('found_enstru'):
                status = "Не найдено"
            else:
                sim = item.get('similarity') or 0
                if sim >= 100:
                    status = "Точное"
                elif sim >= 85:
                    status = "Высокое"
                else:
                    status = "Низкое"

            suppliers_str = ""
            if item.get('suppliers'):
                top = item['suppliers'][:20]
                suppliers_str = "; ".join(top)
                if len(item['suppliers']) > 20:
                    suppliers_str += f" ... и ещё {len(item['suppliers']) - 20}"

            row_data = [
                idx,
                clean_text_for_excel(item.get('found_enstru')),
                clean_text_for_excel(item['name']),
                clean_text_for_excel(item.get('found_name')),
                clean_text_for_excel(item['unit']),
                item['vol'],
                round(price, 2),
                round(item['total'], 2),
                clean_text_for_excel(item['agsk']),
                item['similarity'] if item['similarity'] is not None else "",
                status,
                clean_text_for_excel(suppliers_str)
            ]

            ws.append(row_data)

            # Подсветка строки по качеству
            row_color = self._get_row_color(item)
            fill = PatternFill("solid", fgColor=row_color)
            row_idx = idx + 1
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

        # Фиксируем заголовок
        ws.freeze_panes = "A2"

        filename = f"analysis_{task_id}.xlsx"
        if not os.path.exists(settings.REPORT_DIR):
            os.makedirs(settings.REPORT_DIR)
        path = os.path.join(settings.REPORT_DIR, filename)
        wb.save(path)
        wb.close()

        return filename