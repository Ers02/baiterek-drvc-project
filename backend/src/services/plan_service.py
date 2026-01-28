from sqlalchemy.orm import Session, joinedload, selectinload
from sqlalchemy import func, desc, and_
from decimal import Decimal
from fastapi import HTTPException, status
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import statistics
from ..models import models
from ..schemas import plan as plan_schema

# ========= Вспомогательные функции для версий =========

def _get_active_version(db: Session, plan_id: int, lock: bool = False) -> models.ProcurementPlanVersion | None:
    """Получает активную версию плана."""
    query = db.query(models.ProcurementPlanVersion).filter(
        models.ProcurementPlanVersion.plan_id == plan_id,
        models.ProcurementPlanVersion.is_active == True
    )
    if lock:
        query = query.with_for_update()
    return query.first()

def _recalculate_version_metrics(db: Session, version_id: int):
    """Пересчитывает общую сумму и другие метрики для конкретной версии плана."""
    version = db.query(models.ProcurementPlanVersion).filter(models.ProcurementPlanVersion.id == version_id).first()
    if not version:
        return

    # Получаем все не удаленные позиции
    items = db.query(models.PlanItemVersion).filter(
        models.PlanItemVersion.version_id == version_id,
        models.PlanItemVersion.is_deleted == False
    ).all()

    total_amount = Decimal('0.00')
    vc_amount_total = Decimal('0.00')

    for item in items:
        total_amount += item.total_amount
        
        # Логика определения min_dvc_percent
        if item.need_type == models.NeedType.GOODS:
            # Для товаров ищем в реестре КТП
            min_dvc = db.query(func.min(models.Reestr_KTP.dvc_percent)).filter(
                models.Reestr_KTP.enstru_code == item.trucode
            ).scalar()
            item_dvc_percent = Decimal(str(min_dvc)) if min_dvc is not None else Decimal('0.00')
        else:
            # Для работ и услуг берем из доли местного содержания (resident_share)
            # resident_share уже хранится в модели, используем его
            item_dvc_percent = item.resident_share if item.resident_share is not None else Decimal('0.00')

        item.min_dvc_percent = item_dvc_percent
        
        item_vc_amount = item.total_amount * (item_dvc_percent / Decimal('100.00'))
        item.vc_amount = item_vc_amount # Сохраняем сумму ВЦ в позицию
        
        db.add(item)
        
        vc_amount_total += item_vc_amount

    if total_amount > 0:
        # Доля импорта = (Общая сумма - Сумма ВЦ) / Общая сумма * 100
        import_percentage = ((total_amount - vc_amount_total) / total_amount) * 100
        # Взвешенный процент ВЦ = (Сумма ВЦ / Общая сумма) * 100
        vc_percentage = (vc_amount_total / total_amount) * 100
    else:
        import_percentage = Decimal('0.00')
        vc_percentage = Decimal('0.00')

    version.total_amount = total_amount
    version.import_percentage = import_percentage
    
    version.vc_percentage = vc_percentage
    version.vc_amount = vc_amount_total

    db.commit()
    db.refresh(version)

# ========= Сервисы для Смет Закупок (ProcurementPlan) =========

def create_plan(db: Session, plan_in: plan_schema.ProcurementPlanCreate, user: models.User) -> models.ProcurementPlan:
    db_plan = models.ProcurementPlan(
        plan_name=plan_in.plan_name,
        year=plan_in.year,
        created_by=user.id
    )
    db.add(db_plan)
    db.flush()

    initial_version = models.ProcurementPlanVersion(
        plan_id=db_plan.id,
        version_number=1,
        status=models.PlanStatus.DRAFT,
        is_active=True,
        created_by=user.id
    )
    db.add(initial_version)
    db.commit()
    db.refresh(db_plan)
    return db_plan

def get_plan_with_active_version(db: Session, plan_id: int) -> models.ProcurementPlan | None:
    return db.query(models.ProcurementPlan).options(
        selectinload(models.ProcurementPlan.versions)
        .selectinload(models.ProcurementPlanVersion.items)
        .options(
            joinedload(models.PlanItemVersion.enstru),
            joinedload(models.PlanItemVersion.unit),
            joinedload(models.PlanItemVersion.expense_item),
            joinedload(models.PlanItemVersion.funding_source),
            joinedload(models.PlanItemVersion.agsk),
            joinedload(models.PlanItemVersion.kato_purchase),
            joinedload(models.PlanItemVersion.kato_delivery),
            joinedload(models.PlanItemVersion.source_version),
            joinedload(models.PlanItemVersion.root_item).joinedload(models.PlanItemVersion.version)
        )
    ).filter(
        models.ProcurementPlan.id == plan_id
    ).first()

def get_plans_by_user(db: Session, user: models.User, skip: int = 0, limit: int = 100) -> list[models.ProcurementPlan]:
    return db.query(models.ProcurementPlan).options(
        selectinload(models.ProcurementPlan.versions).selectinload(models.ProcurementPlanVersion.creator)
    ).filter(
        models.ProcurementPlan.created_by == user.id
    ).order_by(desc(models.ProcurementPlan.id)).offset(skip).limit(limit).all()


def update_plan_status(db: Session, plan_id: int, new_status: models.PlanStatus, user: models.User) -> models.ProcurementPlanVersion:
    active_version = _get_active_version(db, plan_id, lock=True)
    if not active_version:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Активная версия плана не найдена")

    current_status = active_version.status

    if current_status == models.PlanStatus.DRAFT and new_status == models.PlanStatus.PRE_APPROVED:
        active_version.status = new_status
    elif current_status == models.PlanStatus.PRE_APPROVED and new_status == models.PlanStatus.APPROVED:
        active_version.status = new_status
    elif current_status == new_status:
        pass
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Недопустимый переход статуса из {current_status.value} в {new_status.value}"
        )

    db.commit()
    db.refresh(active_version)
    return active_version

def create_new_version_for_editing(db: Session, plan_id: int, user: models.User) -> models.ProcurementPlanVersion:
    db.begin_nested()
    try:
        current_active_version = db.query(models.ProcurementPlanVersion).filter(
            models.ProcurementPlanVersion.plan_id == plan_id,
            models.ProcurementPlanVersion.is_active == True
        ).options(
            selectinload(models.ProcurementPlanVersion.items).selectinload(models.PlanItemVersion.executions)
        ).with_for_update().first()

        if not current_active_version:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Активная версия не найдена.")

        if current_active_version.status == models.PlanStatus.DRAFT:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Нельзя создать новую версию из черновика. Сначала одобрите текущую версию.")

        current_active_version.is_active = False
        db.add(current_active_version)

        new_version_number = current_active_version.version_number + 1
        new_version = models.ProcurementPlanVersion(
            plan_id=plan_id,
            version_number=new_version_number,
            status=models.PlanStatus.DRAFT,
            is_active=True,
            created_by=user.id,
            total_amount=current_active_version.total_amount,
            import_percentage=current_active_version.import_percentage
        )
        db.add(new_version)
        db.flush()

        new_items = []
        for item in current_active_version.items:
            if not item.is_deleted:
                new_item_data = {
                    key: getattr(item, key)
                    for key in item.__table__.columns.keys()
                    if key not in ['id', 'version_id', 'created_at']
                }
                new_item_data['version_id'] = new_version.id
                
                new_item_data['root_item_id'] = item.root_item_id if item.root_item_id else item.id
                new_item_data['source_version_id'] = item.source_version_id if item.source_version_id else current_active_version.id
                
                new_items.append(models.PlanItemVersion(**new_item_data))
                
                for execution in item.executions:
                    new_execution_data = {
                        key: getattr(execution, key)
                        for key in execution.__table__.columns.keys()
                        if key not in ['id', 'plan_item_id', 'created_at']
                    }
                    new_execution = models.PlanItemExecution(**new_execution_data)
                    new_execution.plan_item = new_items[-1] 

        if new_items:
            db.add_all(new_items)

        db.commit()
        
        _recalculate_version_metrics(db, new_version.id)
        
        db.refresh(new_version)
        return new_version
    except Exception:
        db.rollback()
        raise

def delete_latest_version(db: Session, plan_id: int, user: models.User):
    db.begin_nested()
    try:
        active_version = _get_active_version(db, plan_id, lock=True)
        if not active_version:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Активная версия не найдена.")

        if active_version.status != models.PlanStatus.DRAFT:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Удалять можно только версию в статусе 'Черновик'.")

        if active_version.version_number == 1:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Нельзя удалить самую первую версию. Вместо этого удалите весь план.")

        previous_version = db.query(models.ProcurementPlanVersion).filter(
            models.ProcurementPlanVersion.plan_id == plan_id,
            models.ProcurementPlanVersion.version_number == active_version.version_number - 1
        ).with_for_update().first()

        if not previous_version:
             raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Предыдущая версия не найдена для восстановления.")

        db.query(models.PlanItemVersion).filter(
            models.PlanItemVersion.version_id == active_version.id
        ).delete(synchronize_session=False)

        db.delete(active_version)

        previous_version.is_active = True
        db.add(previous_version)

        db.commit()
        return {"message": f"Версия {active_version.version_number} удалена. Активной стала версия {previous_version.version_number}."}
    except Exception:
        db.rollback()
        raise

def delete_plan(db: Session, plan_id: int):
    plan_to_delete = db.query(models.ProcurementPlan).options(
        selectinload(models.ProcurementPlan.versions)
    ).filter(models.ProcurementPlan.id == plan_id).first()

    if not plan_to_delete:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="План не найден.")

    has_approved_version = any(
        v.status in [models.PlanStatus.PRE_APPROVED, models.PlanStatus.APPROVED]
        for v in plan_to_delete.versions
    )
    if has_approved_version:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Нельзя удалить план, который уже был одобрен.")

    db.delete(plan_to_delete)
    db.commit()
    return True

# ========= Сервисы для Позиций Плана (PlanItemVersion) =========

def add_item_to_plan(db: Session, plan_id: int, item_in: plan_schema.PlanItemCreate, user: models.User) -> models.PlanItemVersion:
    active_version = _get_active_version(db, plan_id)
    if not active_version:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Активная версия плана не найдена")
    if active_version.status != models.PlanStatus.DRAFT:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Добавлять позиции можно только в черновик.")

    enstru_item = db.query(models.Enstru).filter(models.Enstru.code == item_in.trucode).first()
    if not enstru_item:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Код ЕНС ТРУ не найден")

    # Маппинг type_name на NeedType (Исправлено: добавлена поддержка множественного числа и регистра)
    type_name_upper = enstru_item.type_name.upper() if enstru_item.type_name else 'GOODS'
    
    need_type_map = {
        'GOOD': models.NeedType.GOODS,
        'GOODS': models.NeedType.GOODS,
        'WORK': models.NeedType.WORKS,
        'WORKS': models.NeedType.WORKS,
        'SERVICE': models.NeedType.SERVICES,
        'SERVICES': models.NeedType.SERVICES
    }
    need_type = need_type_map.get(type_name_upper, models.NeedType.GOODS)

    # Ищем последний номер позиции ДЛЯ ЭТОГО ТИПА
    last_item = db.query(models.PlanItemVersion).filter(
        models.PlanItemVersion.version_id == active_version.id,
        models.PlanItemVersion.need_type == need_type # Фильтр по типу
    ).order_by(desc(models.PlanItemVersion.item_number)).first()
    
    item_number = (last_item.item_number + 1) if last_item else 1

    total_amount = item_in.quantity * item_in.price_per_unit

    db_item = models.PlanItemVersion(
        **item_in.model_dump(),
        version_id=active_version.id,
        item_number=item_number,
        total_amount=total_amount,
        need_type=need_type,
        source_version_id=active_version.id,
        revision_number=0
    )
    db.add(db_item)
    db.flush()
    db_item.root_item_id = db_item.id
    db.commit()

    _recalculate_version_metrics(db, active_version.id)

    db.refresh(db_item)
    return db_item

def export_plan_to_excel(db: Session, plan_id: int, version_id: int = None) -> bytes:
    if version_id:
        version = db.query(models.ProcurementPlanVersion).filter(models.ProcurementPlanVersion.id == version_id).first()
    else:
        version = _get_active_version(db, plan_id)

    if not version:
        raise HTTPException(status_code=404, detail="Версия сметы не найдена")

    # Загружаем версию вместе с планом и создателем плана (для получения наименования клиента)
    version_with_items = db.query(models.ProcurementPlanVersion).options(
        selectinload(models.ProcurementPlanVersion.items).options(
            joinedload(models.PlanItemVersion.enstru),
            joinedload(models.PlanItemVersion.unit),
            joinedload(models.PlanItemVersion.expense_item),
            joinedload(models.PlanItemVersion.funding_source),
            joinedload(models.PlanItemVersion.agsk),
            joinedload(models.PlanItemVersion.kato_purchase),
            joinedload(models.PlanItemVersion.kato_delivery),
            joinedload(models.PlanItemVersion.source_version),
            joinedload(models.PlanItemVersion.root_item).joinedload(models.PlanItemVersion.version)
        ),
        joinedload(models.ProcurementPlanVersion.plan).joinedload(models.ProcurementPlan.creator)
    ).filter(models.ProcurementPlanVersion.id == version.id).one()

    wb = openpyxl.Workbook()
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    sub_header_font = Font(bold=True, color="000000")
    sub_header_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    def format_item_number(idx, item):
        # Используем переданный индекс для последовательной нумерации
        number = f"{idx}"
        if item.revision_number > 0:
            number += f"-{item.revision_number}"
        
        type_suffix = ""
        if item.need_type == models.NeedType.GOODS: type_suffix = " Т"
        elif item.need_type == models.NeedType.WORKS: type_suffix = " Р"
        elif item.need_type == models.NeedType.SERVICES: type_suffix = " У"
        
        return f"{number}{type_suffix}"

    grouped_items = {
        models.NeedType.GOODS: [],
        models.NeedType.WORKS: [],
        models.NeedType.SERVICES: []
    }
    
    # Сортировка по item_number внутри групп
    for item in version_with_items.items:
        if not item.is_deleted:
            grouped_items[item.need_type].append(item)
            
    for key in grouped_items:
        grouped_items[key].sort(key=lambda x: x.item_number)

    # --- Лист 1: Основная смета ---
    ws = wb.active
    ws.title = "Смета"
    
    # Заголовок и информация о проекте
    ws.merge_cells('A1:S1')
    ws['A1'] = "СМЕТА ЗАКУПОК"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Наименование проекта
    ws.merge_cells('A2:S2')
    ws['A2'] = f"Наименование проекта: {version_with_items.plan.plan_name}"
    ws['A2'].font = Font(bold=True, size=12)
    
    # Год
    ws.merge_cells('A3:S3')
    ws['A3'] = f"Год: {version_with_items.plan.year}"
    ws['A3'].font = Font(bold=True, size=12)
    
    # Наименование клиента
    client_name = version_with_items.plan.creator.org_name if version_with_items.plan.creator and version_with_items.plan.creator.org_name else "Не указано"
    ws.merge_cells('A4:S4')
    ws['A4'] = f"Наименование клиента: {client_name}"
    ws['A4'].font = Font(bold=True, size=12)
    
    current_row = 6
    
    columns = [
        "№", 
        "Код по ЕНС ТРУ", 
        "Наименование закупаемых товаров услуг работ (рус)", 
        "Наименование закупаемых товаров услуг работ (каз)", 
        "Краткая характеристика (рус)",
        "Краткая характеристика (каз)",
        "Дополнительная характеристика (рус)",
        "Дополнительная характеристика (каз)",
        "Единица измерения(МКЕИ)",
        "Колво объём",
        "цена за единицу тенге",
        "сумма планируемая для закупок ТРУ",
        "Место закупки", 
        "Место поставки",
        "Статья затрат",
        "источник финансирования",
        "КОД АГСК для смр",
        "КТП",
        "ВЦ %",
        "Сумма ВЦ тенге без НДС"
    ]
    
    def create_table_header(ws, row_idx, cols):
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[row_idx].height = 45
        return row_idx + 1

    def fill_section(title, items, start_row):
        if not items: return start_row
        
        ws.merge_cells(f'A{start_row}:S{start_row}')
        ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12)
        ws.cell(row=start_row, column=1).fill = sub_header_fill
        start_row += 1
        
        start_row = create_table_header(ws, start_row, columns)
        
        section_total = Decimal('0.00')
        section_vc_amount = Decimal('0.00')
        
        for idx, item in enumerate(items, 1):
            # Логика для АГСК: если СМР и agsk_id нет, то "Прайс-лист"
            agsk_value = ""
            if item.expense_item and item.expense_item.name_ru == "СМР":
                if item.agsk_id:
                    agsk_value = item.agsk_id
                else:
                    agsk_value = "Прайс-лист"
            elif item.agsk_id:
                agsk_value = item.agsk_id

            row_data = [
                format_item_number(idx, item), # Передаем порядковый номер
                item.trucode,
                item.enstru.name_rus if item.enstru else "",
                item.enstru.name_kaz if item.enstru else "",
                item.enstru.detail_rus if item.enstru else "",
                item.enstru.detail_kaz if item.enstru else "",
                item.additional_specs,
                item.additional_specs_kz,
                item.unit.name_ru if item.unit else "",
                item.quantity,
                item.price_per_unit,
                item.total_amount,
                item.kato_purchase.name_ru if item.kato_purchase else "",
                item.kato_delivery.name_ru if item.kato_delivery else "",
                item.expense_item.name_ru if item.expense_item else "",
                item.funding_source.name_ru if item.funding_source else "",
                agsk_value, # Используем вычисленное значение
                "Да" if item.is_ktp else "Нет",
                f"{item.min_dvc_percent}",
                item.vc_amount
            ]
            
            section_total += item.total_amount
            section_vc_amount += item.vc_amount
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=val)
                cell.border = border
                if col_idx in [10, 11, 12, 20]: # Числовые поля
                    cell.number_format = '#,##0.00'
            
            start_row += 1
            
        # Итого по разделу
        ws.merge_cells(f'A{start_row}:K{start_row}')
        ws.cell(row=start_row, column=1, value=f"Итого по {title.lower()}:").font = Font(bold=True)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='right')
        ws.cell(row=start_row, column=12, value=section_total).font = Font(bold=True)
        ws.cell(row=start_row, column=12).number_format = '#,##0.00'
        
        # Добавляем итоги по ВЦ для раздела (взвешенное среднее)
        section_vc_mean = (section_vc_amount / section_total * 100) if section_total > 0 else Decimal('0.00')
        
        ws.cell(row=start_row, column=19, value=f"Ср. {section_vc_mean.quantize(Decimal('0.00'))}%").font = Font(bold=True)
        ws.cell(row=start_row, column=20, value=section_vc_amount).font = Font(bold=True)
        ws.cell(row=start_row, column=20).number_format = '#,##0.00'
        
        return start_row + 2

    current_row = fill_section("1. Товары", grouped_items[models.NeedType.GOODS], current_row)
    current_row = fill_section("2. Работы", grouped_items[models.NeedType.WORKS], current_row)
    current_row = fill_section("3. Услуги", grouped_items[models.NeedType.SERVICES], current_row)

    # Всего
    ws.merge_cells(f'A{current_row}:K{current_row}')
    ws.cell(row=current_row, column=1, value="Всего:").font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right')
    ws.cell(row=current_row, column=12, value=version_with_items.total_amount).font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=12).number_format = '#,##0.00'
    
    # Расчет общего взвешенного среднего процента ВЦ
    total_vc_mean = (version_with_items.vc_amount / version_with_items.total_amount * 100) if version_with_items.total_amount > 0 else Decimal('0.00')

    current_row += 1
    ws.cell(row=current_row, column=9, value="Средний % ВЦ:").font = Font(bold=True)
    ws.cell(row=current_row, column=10, value=f"{total_vc_mean.quantize(Decimal('0.00'))}%").font = Font(bold=True)
    
    current_row += 1
    ws.cell(row=current_row, column=9, value="Общая сумма ВЦ:").font = Font(bold=True)
    ws.cell(row=current_row, column=10, value=version_with_items.vc_amount).font = Font(bold=True)
    ws.cell(row=current_row, column=10).number_format = '#,##0.00'
    
    # Автоширина колонок
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)


    # --- Лист 2: КТП ---
    ws_ktp = wb.create_sheet("КТП")
    
    ktp_columns = [
        "№", 
        "Код по ЕНС ТРУ", 
        "Наименование закупаемых товаров услуг работ", 
        "Краткая характеристика",
        "Дополнительная характеристика (рус)",
        "Дополнительная характеристика (каз)",
        "Единица измерения(МКЕИ)",
        "Колво объём",
        "цена за единицу тенге",
        "сумма планируемая для закупок ТРУ",
        "Место закупки", 
        "Место поставки",
        "Статья затрат",
        "источник финансирования",
        "КОД АГСК для смр",
        "КТП",
        "Сумма ВЦ тенге без НДС",
        "БИН производителя",
        "Наименования производителя",
        "Адрес/ контакты",
        "ВЦ% по этому производителю",
        "Сумма ВЦ тенге без НДС (по производителю)"
    ]
    
    ktp_row = 1
    ktp_row = create_table_header(ws_ktp, ktp_row, ktp_columns)
    
    for t in [models.NeedType.GOODS, models.NeedType.WORKS, models.NeedType.SERVICES]:
        items = grouped_items[t]
        for idx, item in enumerate(items, 1):
            # Проверяем наличие в реестре КТП
            suppliers = db.query(models.Reestr_KTP).filter(models.Reestr_KTP.enstru_code == item.trucode).all()
            
            if suppliers:
                # Для каждого поставщика создаем строку
                for supplier in suppliers:
                    supplier_dvc = Decimal(str(supplier.dvc_percent)) if supplier.dvc_percent is not None else Decimal('0.00')
                    supplier_vc_amount = item.total_amount * (supplier_dvc / Decimal('100.00'))
                    
                    # Логика для АГСК (дублируем)
                    agsk_value = ""
                    if item.expense_item and item.expense_item.name_ru == "СМР":
                        if item.agsk_id:
                            agsk_value = item.agsk_id
                        else:
                            agsk_value = "Прайс-лист"
                    elif item.agsk_id:
                        agsk_value = item.agsk_id

                    row_data = [
                        format_item_number(idx, item), # Используем индекс позиции
                        item.trucode,
                        item.enstru.name_rus if item.enstru else "",
                        item.enstru.detail_rus if item.enstru else "",
                        item.additional_specs,
                        item.additional_specs_kz,
                        item.unit.name_ru if item.unit else "",
                        item.quantity,
                        item.price_per_unit,
                        item.total_amount,
                        item.kato_purchase.name_ru if item.kato_purchase else "",
                        item.kato_delivery.name_ru if item.kato_delivery else "",
                        item.expense_item.name_ru if item.expense_item else "",
                        item.funding_source.name_ru if item.funding_source else "",
                        agsk_value,
                        "Да" if item.is_ktp else "Нет",
                        item.vc_amount, # Сумма ВЦ общая (по мин. проценту)
                        
                        supplier.bin_iin,
                        supplier.company_name,
                        f"{supplier.production_address or ''} {supplier.phone or ''} {supplier.email or ''}",
                        f"{supplier_dvc}",
                        supplier_vc_amount
                    ]
                    
                    for col_idx, val in enumerate(row_data, 1):
                        cell = ws_ktp.cell(row=ktp_row, column=col_idx, value=val)
                        cell.border = border
                        if col_idx in [8, 9, 10, 17, 22]:
                            cell.number_format = '#,##0.00'
                    
                    ktp_row += 1

    # Автоширина для КТП
    for i, col in enumerate(ws_ktp.columns, 1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_ktp.column_dimensions[column_letter].width = min(adjusted_width, 50)

    virtual_workbook = io.BytesIO()
    wb.save(virtual_workbook)
    return virtual_workbook.getvalue()
