from decimal import Decimal
from typing import Optional
from pydantic import BaseModel, Field, ValidationError, field_validator
from sqlalchemy.orm import Session
from sqlalchemy import desc, func
import openpyxl
import io
from ...models import models
from ...utils.helpers import get_need_type_by_typename, is_smr
from ..dictionary_service import get_mkei_map, get_cost_item_map, get_kato_map, get_agsk_map


class ImportRow(BaseModel):
    row_idx: int
    trucode: str
    additional_specs: str
    unit_raw: Optional[str] = None
    quantity: Decimal = Field(gt=0)
    price: Decimal = Field(ge=0)
    kato_purchase_code: str
    kato_delivery_code: str
    expense_id: int
    source_id: int
    agsk_code: Optional[str] = None
    resident_share: Decimal = Field(default=100, ge=0, le=100)
    non_resident_reason: Optional[str] = None

    @field_validator('trucode')
    def validate_trucode(cls, v):
        if not v or not v.strip():
            raise ValueError("Код ЕНС ТРУ обязателен")
        return v.strip()


def extract_code(val):
    if val is None: return None
    val_str = str(val).strip()
    if not val_str: return None
    if " - " in val_str:
        return val_str.split(" - ")[0].strip()
    return val_str


def parse_excel_rows(ws) -> list[dict]:
    """Читает Excel и возвращает список сырых словарей."""
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
            
        row_data = list(row) + [None] * max(0, 16 - len(row))
        
        try:
            qty = row_data[5]
            price = row_data[6]
            res_share = row_data[13]
            if res_share is None: res_share = 100

            data = {
                "row_idx": row_idx,
                "trucode": str(row_data[1]) if row_data[1] else "",
                "additional_specs": str(row_data[3]).strip() if row_data[3] else "",
                "unit_raw": str(row_data[4]).strip() if row_data[4] else None,
                "quantity": qty,
                "price": price,
                "kato_purchase_code": extract_code(row_data[8]) or "",
                "kato_delivery_code": extract_code(row_data[9]) or "",
                "expense_id": int(extract_code(row_data[10]) or 0),
                "source_id": int(extract_code(row_data[11]) or 0),
                "agsk_code": str(row_data[12]).strip() if (row_data[12] and str(row_data[12]).strip()) else None,
                "resident_share": res_share,
                "non_resident_reason": str(row_data[14]).strip() if row_data[14] else None
            }
            rows.append(data)
        except Exception as e:
            rows.append({"row_idx": row_idx, "error": str(e)})
            
    return rows


def import_items_from_excel(db: Session, version_id: int, file_content: bytes) -> tuple[list[models.PlanItemVersion], list[dict]]:
    # RE-APPLY FIX: Updated logic for KTP validation with JSONB
    """
    Основная функция импорта.
    Возвращает (список созданных объектов, список ошибок).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb["Позиции для загрузки"] if "Позиции для загрузки" in wb.sheetnames else wb.active
    
    raw_rows = parse_excel_rows(ws)
    valid_rows = []
    errors = []

    trucodes_to_fetch = set()

    for r in raw_rows:
        if "error" in r:
            errors.append({"row": r["row_idx"], "message": f"Ошибка чтения строки: {r['error']}"})
            continue
            
        try:
            row_obj = ImportRow(**r)
            valid_rows.append(row_obj)
            trucodes_to_fetch.add(row_obj.trucode)
        except ValidationError as e:
            msg = "; ".join([f"{err['loc'][0]}: {err['msg']}" for err in e.errors()])
            errors.append({"row": r["row_idx"], "message": msg})

    mkei_map = get_mkei_map()
    kato_map = get_kato_map()
    agsk_map = get_agsk_map()
    cost_map = get_cost_item_map()
    
    enstru_map = {}
    if trucodes_to_fetch:
        items = db.query(models.Enstru.code, models.Enstru.type_name).filter(models.Enstru.code.in_(trucodes_to_fetch)).all()
        for code, type_name in items:
            enstru_map[code] = type_name
            
    # UPDATED: Validation logic for KTP using new JSONB structure
    reestr_ktp_map = {}
    if trucodes_to_fetch:
        # Fetch all potential candidates (optimize by checking if enstru_codes is not null)
        ktp_candidates = db.query(
            models.Reestr_KTP.enstru_codes,
            models.Reestr_KTP.dvc_percent
        ).filter(
            models.Reestr_KTP.enstru_codes.isnot(None),
            models.Reestr_KTP.dvc_percent.isnot(None)
        ).all()

        for candidate in ktp_candidates:
            codes = candidate.enstru_codes
            dvc_str = candidate.dvc_percent
            
            if not codes or not dvc_str:
                continue
                
            try:
                # Handle text format like "50,5" or "50.5" or just "50"
                dvc = Decimal(str(dvc_str).replace(',', '.'))
            except Exception:
                continue
                
            if dvc <= 0:
                continue

            # Check if any of the candidate's codes match our needed codes
            # Intersection is faster than iterating
            common_codes = set(codes).intersection(trucodes_to_fetch)
            
            for code in common_codes:
                if code not in reestr_ktp_map:
                    reestr_ktp_map[code] = dvc
                else:
                    # Keep the minimum DVC found
                    if dvc < reestr_ktp_map[code]:
                        reestr_ktp_map[code] = dvc

    new_items = []
    
    last_numbers = {
        models.NeedType.GOODS: 0,
        models.NeedType.WORKS: 0,
        models.NeedType.SERVICES: 0
    }
    for nt in last_numbers:
        last = db.query(models.PlanItemVersion).filter(
            models.PlanItemVersion.version_id == version_id,
            models.PlanItemVersion.need_type == nt
        ).order_by(desc(models.PlanItemVersion.item_number)).first()
        if last: last_numbers[nt] = last.item_number

    for row in valid_rows:
        if row.trucode not in enstru_map:
            errors.append({"row": row.row_idx, "message": f"Не найден ЕНС ТРУ {row.trucode}"})
            continue
            
        kato_p_id = kato_map.get(row.kato_purchase_code)
        kato_d_id = kato_map.get(row.kato_delivery_code)
        if not kato_p_id or not kato_d_id:
            errors.append({"row": row.row_idx, "message": "Не найден КАТО"})
            continue
            
        if row.expense_id not in cost_map:
             errors.append({"row": row.row_idx, "message": f"Не найдена статья затрат {row.expense_id}"})
             continue

        is_smr_item = is_smr(row.expense_id)
        
        agsk_code = None
        if row.agsk_code:
            if row.agsk_code.lower() == "прайс-лист":
                agsk_code = None
            else:
                if row.agsk_code not in agsk_map:
                    errors.append({"row": row.row_idx, "message": f"Не найден АГСК {row.agsk_code}"})
                    continue
                agsk_code = row.agsk_code
        
        if is_smr_item and agsk_code is None and (not row.agsk_code or row.agsk_code.lower() != "прайс-лист"):
             errors.append({"row": row.row_idx, "message": "Для СМР (ID=1) нужен код АГСК или 'Прайс-лист'"})
             continue

        type_name = enstru_map[row.trucode]
        need_type = get_need_type_by_typename(type_name)
        
        unit_id = None
        original_unit_name = None
        
        if is_smr_item:
            original_unit_name = row.unit_raw
        elif need_type == models.NeedType.GOODS:
            unit_code = extract_code(row.unit_raw)
            if unit_code and unit_code in mkei_map:
                unit_id = mkei_map[unit_code]
            else:
                original_unit_name = row.unit_raw
            
            if not unit_id and not original_unit_name:
                errors.append({"row": row.row_idx, "message": "Не указана единица измерения"})
                continue
        
        quantity = row.quantity
        if need_type != models.NeedType.GOODS:
            quantity = Decimal(1)
            
        resident_share = row.resident_share
        non_resident_reason = row.non_resident_reason
        if need_type == models.NeedType.GOODS:
            resident_share = Decimal(0)
            non_resident_reason = None
        else:
            if resident_share < 100 and not non_resident_reason:
                errors.append({"row": row.row_idx, "message": "Нужно обоснование для доли < 100%"})
                continue

        is_ktp = False
        min_dvc = Decimal(0)
        if need_type == models.NeedType.GOODS:
            dvc = reestr_ktp_map.get(row.trucode)
            if dvc is not None:
                is_ktp = True
                min_dvc = dvc
        else:
            min_dvc = resident_share

        last_numbers[need_type] += 1
        
        item = models.PlanItemVersion(
            version_id=version_id,
            item_number=last_numbers[need_type],
            need_type=need_type,
            trucode=row.trucode,
            unit_id=unit_id,
            original_unit_name=original_unit_name,
            expense_item_id=row.expense_id,
            funding_source_id=row.source_id,
            agsk_code=agsk_code,
            kato_purchase_id=kato_p_id,
            kato_delivery_id=kato_d_id,
            additional_specs=row.additional_specs,
            quantity=quantity,
            price_per_unit=row.price,
            total_amount=quantity * row.price,
            is_ktp=is_ktp,
            is_deleted=False,
            revision_number=0,
            min_dvc_percent=min_dvc,
            resident_share=resident_share,
            non_resident_reason=non_resident_reason
        )
        new_items.append(item)

    return new_items, errors
